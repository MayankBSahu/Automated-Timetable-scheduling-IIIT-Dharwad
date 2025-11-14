
import os
import math
import pandas as pd
import random
import hashlib
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

Random_SEED = 314156
random.seed(Random_SEED)

# --------------------- Utility functions ---------------------
def stable_hash_val(x: object) -> int:
    """Return a stable integer hash for object x using SHA-256 (deterministic across runs)."""
    h = hashlib.sha256(str(x).encode("utf-8")).hexdigest()
    return int(h[:16], 16)


def stable_key(x: object) -> int:
    """Combine stable hash and global Random_SEED to provide deterministic ordering."""
    return stable_hash_val(x) ^ Random_SEED


# --------------------- Data container ---------------------
class Course:
    """Container for course attributes (code, title, L-T-P-S-C, faculty, basket, elective flag)."""

    def __init__(self, row):
        self.code = str(row["Course_Code"]).strip()
        self.basket = int(row.get("basket", 0))
        self.title = str(row.get("Course_Title", self.code)).strip()
        self.faculty = str(row.get("Faculty", "")).strip()
        self.ltp = str(row["L-T-P-S-C"]).strip()
        # keep original CSV field name semantics
        self.sem_half = str(row.get("Semester_Half", "0")).strip()
        self.is_elective = str(row.get("Elective", 0)).strip() == "1"
        # parse L-T-P-S-C; be tolerant to malformed values
        try:
            parts = list(map(int, self.ltp.split("-")))
            # some CSVs have 5 parts, some 3; ensure at least L, T, P present
            if len(parts) >= 5:
                self.L, self.T, self.P, self.S, self.C = parts[:5]
            else:
                # pad shorter arrays
                while len(parts) < 5:
                    parts.append(0)
                self.L, self.T, self.P, self.S, self.C = parts[:5]
        except Exception:
            self.L, self.T, self.P, self.S, self.C = 0, 0, 0, 0, 0


# --------------------- Scheduler ---------------------
class Scheduler:
    """
    Responsible for:
    - Reading slots, courses, rooms
    - Building timetables (First_Half and Second_Half)
    - Assigning rooms (lab vs classroom)
    - Applying deterministic tie-breaking (stable hash + seed)
    - Exporting timetables and faculty sheets to Excel and applying formatting
    """

    def __init__(self, slots_file, courses_file, rooms_file, global_room_usage):
        # Read timeslots
        slot_frame = pd.read_csv(slots_file)
        self.slots = [f"{r['Start_Time'].strip()}-{r['End_Time'].strip()}" for _, r in slot_frame.iterrows()]
        self.slot_lengths = {s: self._slot_len(s) for s in self.slots}

        # Read courses
        course_data = pd.read_csv(courses_file)
        self.courses = [Course(row) for _, row in course_data.iterrows()]

        # Read rooms
        rooms_df = pd.read_csv(rooms_file)
        self.classrooms, self.labs, self.all_rooms = [], [], []
        for _, row in rooms_df.iterrows():
            room_id = str(row["Room_ID"]).strip()
            self.all_rooms.append(room_id)
            if room_id.upper().startswith("L"):
                self.labs.append(room_id)
            elif room_id.upper().startswith("C"):
                self.classrooms.append(room_id)

        # Scheduling parameters
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.excluded = ["07:30-09:00", "13:15-14:00", "17:40-18:30"]
        self.MAX_ATTEMPTS = 10

        # Mutable state that gets populated during scheduling
        self.unscheduled_list = []
        self.course_room_map = {}
        self.global_room_usage = global_room_usage  # external mapping across depts
        self.records = []  # each scheduled placement as dict
        self.elective_groups = {}
        self.elective_room_map = {}
        self.break_after_slots = 1

    # --------------------- Helpers ---------------------
    def _slot_len(self, slot):
        start, end = slot.split("-")
        h1, m1 = map(int, start.split(":"))
        h2, m2 = map(int, end.split(":"))
        return (h2 + m2 / 60) - (h1 + m1 / 60)

    def _free_blocks(self, table, day):
        """
        Return contiguous free-slot blocks (lists of slot keys) for a given day's timetable.
        Excluded slots are treated as occupied.
        """
        blocks, tmp = [], []
        for slot in self.slots:
            if table.at[day, slot] == "" and slot not in self.excluded:
                tmp.append(slot)
            else:
                if tmp:
                    blocks.append(tmp)
                    tmp = []
        if tmp:
            blocks.append(tmp)
        return blocks

    # --------------------- Core allocation ---------------------
    def _assign_session(self, table, faculty_busy, lab_flag, day, faculty, code, hrs, session_type="L", is_elective=False, sheet_name=None):
        """
        Try to place a contiguous session of 'hrs' hours on 'day' for course 'code'.
        Respects faculty busy times, lab-day restrictions, room availability, and excluded slots.
        Adds entries to self.records and updates global_room_usage when room assigned.
        Returns True on successful placement.
        """
        # prevent placing same course multiple times on same day for same sheet
        for rec in self.records:
            if rec["day"] == day and rec["code"] == code and rec["sheet"] == sheet_name:
                return False

        # cannot schedule practical if a lab already scheduled that day (policy from original code)
        if session_type == "P" and lab_flag[day]:
            return False

        free_blocks = self._free_blocks(table, day)
        for block in free_blocks:
            total = sum(self.slot_lengths[s] for s in block)
            if total >= hrs:
                slots_to_use, dur_accum = [], 0.0
                for s in block:
                    slots_to_use.append(s)
                    dur_accum += self.slot_lengths[s]
                    if dur_accum >= hrs:
                        break

                # faculty availability check
                if faculty:
                    busy = any(faculty in faculty_busy[day][s] for s in slots_to_use)
                    if busy:
                        continue

                # room assignment
                if not is_elective:
                    mapped = self.course_room_map.get(code)
                    if mapped:
                        # if mapped room suits session type, use it
                        if (session_type == "P" and mapped.upper().startswith("L")) or (session_type != "P" and not mapped.upper().startswith("L")):
                            room = mapped
                        else:
                            mapped = None
                    if not mapped:
                        possible_rooms = self.labs if session_type == "P" else self.classrooms
                        available_rooms = [
                            r for r in possible_rooms
                            if all(r not in self.global_room_usage.get(day, {}).get(s, []) for s in slots_to_use)
                        ]
                        if not available_rooms:
                            return False
                        # deterministic selection using stable ordering and numeric seed
                        room = sorted(available_rooms, key=lambda r: stable_key(r))[Random_SEED % len(available_rooms)]
                        self.course_room_map[code] = room

                    # mark room usage
                    for s in slots_to_use:
                        self.global_room_usage.setdefault(day, {}).setdefault(s, []).append(room)
                else:
                    room = ""

                # write to timetable and records
                for i, s in enumerate(slots_to_use):
                    if session_type == "L":
                        display_text = f"{code} ({room})" if (room and not is_elective) else code
                    elif session_type == "T":
                        display_text = f"{code}T ({room})" if (room and not is_elective) else f"{code}T"
                    elif session_type == "P":
                        display_text = f"{code} (Lab-{room})" if (room and not is_elective) else code
                    else:
                        display_text = code

                    table.at[day, s] = display_text
                    self.records.append({
                        "sheet": sheet_name,
                        "day": day,
                        "slot": s,
                        "code": code,
                        "display": display_text,
                        "faculty": faculty,
                        "room": room,
                    })

                    # prevent tiny-gap double booking for quarter-hour small breaks
                    if i < len(slots_to_use) - 1:
                        idx = self.slots.index(s)
                        if idx + 1 < len(self.slots):
                            gap_slot = self.slots[idx + 1]
                            if table.at[day, gap_slot] == "" and math.isclose(self.slot_lengths[gap_slot], 0.25):
                                table.at[day, gap_slot] = "FREE"

                # mark faculty busy
                if faculty:
                    for s in slots_to_use:
                        faculty_busy[day][s].append(faculty)

                # flag that a lab was scheduled that day
                if session_type == "P":
                    lab_flag[day] = True

                # insert post-session break slots (if empty)
                last_slot = slots_to_use[-1]
                idx = self.slots.index(last_slot)
                for extra in range(1, self.break_after_slots + 1):
                    if idx + extra < len(self.slots):
                        next_slot = self.slots[idx + extra]
                        if table.at[day, next_slot] == "":
                            table.at[day, next_slot] = "BREAK"
                            if faculty:
                                faculty_busy[day][next_slot].append(faculty)
                            if not is_elective and room:
                                self.global_room_usage.setdefault(day, {}).setdefault(next_slot, []).append(room)

                return True

        return False

    # --------------------- Timetable generation ---------------------
    def generate_timetable(self, course_list, writer, sheet_name):
        """
        Build a timetable DataFrame for the given course_list and write it to the provided Excel writer
        under 'sheet_name'. This fills self.records and possibly self.unscheduled_list.
        """
        timetable = pd.DataFrame("", index=self.days, columns=self.slots)
        faculty_busy = {day: {slot: [] for slot in self.slots} for day in self.days}
        labs_scheduled = {day: False for day in self.days}
        self.course_room_map = {}

        # separate electives and non-electives
        electives = [c for c in course_list if c.is_elective]
        non_electives = [c for c in course_list if not c.is_elective]

        # group electives by basket
        baskets = {}
        for e in electives:
            baskets.setdefault(e.basket, []).append(e)

        chosen_electives = []
        # choose one elective per basket deterministically
        for b in sorted(baskets.keys()):
            if b == 0:
                continue
            group = baskets[b]
            pick = group[stable_hash_val(b) % len(group)]
            chosen_electives.append((b, pick))

            # create a placeholder course entry for the chosen elective (so it gets scheduled like a normal course)
            elective_course = Course({
                "Course_Code": f"Elective_{b}",
                "Course_Title": pick.title,
                "Faculty": pick.faculty,
                "L-T-P-S-C": pick.ltp,
                "Semester_Half": pick.sem_half,
                "Elective": 0,
            })
            non_electives.append(elective_course)

        self.elective_groups[sheet_name] = chosen_electives

        # deterministic ordering of non-electives (stable_key ensures constant ordering)
        non_electives.sort(key=lambda c: stable_key(c.code))

        # allocate for each course: Lectures (L), Tutorials (T), Practicals (P)
        for course in non_electives:
            faculty, code, is_elective = course.faculty, course.code, course.code.startswith("Elective_")

            # Lectures (each lecture block may be 1.5 hours)
            remaining, attempts = course.L, 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = sorted(self.days, key=lambda d: stable_key(f"{d}-L"))
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in faculty_busy[day]):
                        continue
                    alloc = min(1.5, remaining)
                    if self._assign_session(timetable, faculty_busy, labs_scheduled, day, faculty, code, alloc, "L", is_elective, sheet_name):
                        remaining -= alloc
                        break

            if remaining > 0:
                self.unscheduled_list.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lecture",
                    "remaining_hours": remaining,
                    "semester_half": course.sem_half
                })

            # Tutorials (1 hour each)
            remaining, attempts = course.T, 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = sorted(self.days, key=lambda d: stable_key(f"{d}-T"))
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in faculty_busy[day]):
                        continue
                    if self._assign_session(timetable, faculty_busy, labs_scheduled, day, faculty, code, 1, "T", is_elective, sheet_name):
                        remaining -= 1
                        break

            if remaining > 0:
                self.unscheduled_list.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Tutorial",
                    "remaining_hours": remaining,
                    "semester_half": course.sem_half
                })

            # Practicals (labs)
            remaining, attempts = course.P, 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_without_labs = [d for d in self.days if not labs_scheduled[d]]
                days_to_try = sorted(days_without_labs, key=lambda d: stable_key(f"{d}-P"))
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in faculty_busy[day]):
                        continue
                    alloc = 2 if remaining >= 2 else remaining
                    if self._assign_session(timetable, faculty_busy, labs_scheduled, day, faculty, code, alloc, "P", is_elective, sheet_name):
                        remaining -= alloc
                        break

            if remaining > 0:
                self.unscheduled_list.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lab",
                    "remaining_hours": remaining,
                    "semester_half": course.sem_half
                })

        # Clear excluded slots in final timetable (set to empty string)
        for day in self.days:
            for slot in self.excluded:
                if slot in timetable.columns:
                    timetable.at[day, slot] = ""

        # Write timetable sheet
        timetable.to_excel(writer, sheet_name=sheet_name, index=True)
        print(f"Saved timetable sheet: {sheet_name}")

    # --------------------- Elective room assignment ---------------------
    def _compute_elective_room_assignments_legally(self, sheet_name):
        """
        Assign rooms to chosen elective placeholders ensuring no room conflicts.
        This produces a mapping self.elective_room_map[sheet_name] = {key: room}
        where key is 'Elective_basket||Title'.
        """
        electives = self.elective_groups.get(sheet_name, [])
        if not electives:
            self.elective_room_map[sheet_name] = {}
            return

        # gather (day, slot) pairs where elective placeholders are scheduled
        elective_slots = [(r["day"], r["slot"]) for r in self.records if r["sheet"] == sheet_name and r["code"].startswith("Elective_")]
        elective_slots = sorted(list(dict.fromkeys(elective_slots)))

        candidate_rooms = list(self.classrooms) + list(self.labs)
        room_free_all_slots = {}
        for r in candidate_rooms:
            ok = True
            for day, slot in elective_slots:
                if r in self.global_room_usage.get(day, {}).get(slot, []):
                    ok = False
                    break
            room_free_all_slots[r] = ok

        free_rooms = [r for r, ok in room_free_all_slots.items() if ok]
        if not free_rooms:
            # compute counts of free slots per room and sort by best fit
            room_free_counts = {r: 0 for r in candidate_rooms}
            for r in candidate_rooms:
                for day, slot in elective_slots:
                    if r not in self.global_room_usage.get(day, {}).get(slot, []):
                        room_free_counts[r] += 1
            ordered = sorted(candidate_rooms, key=lambda x: (-room_free_counts[x], stable_key(x)))
            free_rooms = ordered

        assigned = {}
        used = set()
        idx = 0
        for basket, elective in electives:
            key = f"Elective_{basket}||{elective.title}"
            chosen_room = None
            for r in free_rooms:
                if r in used:
                    continue
                ok = True
                for day, slot in elective_slots:
                    if r in self.global_room_usage.get(day, {}).get(slot, []):
                        ok = False
                        break
                if ok:
                    chosen_room = r
                    break
            if not chosen_room:
                chosen_room = free_rooms[idx % len(free_rooms)] if free_rooms else ""
            assigned[key] = chosen_room
            used.add(chosen_room)
            idx += 1

        self.elective_room_map[sheet_name] = assigned

    # --------------------- Excel formatting for student timetables ---------------------
    def format_student_timetable_with_legend(self, filename):
        """
        Apply coloring, merges, borders and a legend to the student timetable workbook.
        A legend is added below each sheet summarizing course codes, titles and colors.
        """
        wb = load_workbook(filename)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        color_map = {}
        palette = ["FFC7CE", "C6EFCE", "FFEB9C", "BDD7EE", "D9EAD3", "F4CCCC", "D9D2E9", "FCE5CD", "C9DAF8", "EAD1DC"]
        color_index = 0

        # precompute elective room assignments
        for sheet_name in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet_name)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # color and merge cells for each row
            for row in range(2, ws.max_row + 1):
                start_col = 2
                while start_col <= ws.max_column:
                    cell = ws.cell(row=row, column=start_col)
                    val = str(cell.value).strip() if cell.value is not None else ""

                    if val and val not in ["FREE", "BREAK"]:
                        raw_code = val.split(" ")[0]
                        code = raw_code.rstrip("T")
                        if code not in color_map:
                            color_map[code] = palette[color_index % len(palette)]
                            color_index += 1
                        fill = PatternFill(start_color=color_map[code], end_color=color_map[code], fill_type="solid")

                        cell.fill = fill

                        merge_count = 0
                        for col in range(start_col + 1, ws.max_column + 1):
                            next_cell = ws.cell(row=row, column=col)
                            if next_cell.value == cell.value:
                                next_cell.fill = fill
                                merge_count += 1
                            else:
                                break
                        if merge_count > 0:
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + merge_count)

                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(start_col, start_col + merge_count + 1):
                            ws.cell(row=row, column=col_idx).border = thin_border
                        start_col += merge_count + 1
                    elif val == "BREAK":
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        start_col += 1
                    else:
                        cell.border = thin_border
                        start_col += 1

            # add legend header and entries (student timetable legend)
            start_row = ws.max_row + 3
            headers = ["S.No", "Course Code", "Course Title", "L-T-P-S-C", "Faculty", "Color"]
            for idx, header in enumerate(headers, start=2):
                ws.cell(start_row, idx, header).border = thin_border
                ws.cell(start_row, idx).alignment = Alignment(horizontal="center", vertical="center")

            i = 1
            for code in color_map:
                if code.startswith("Elective_"):
                    continue
                ws.cell(start_row + i, 2, i).border = thin_border
                ws.cell(start_row + i, 3, code).border = thin_border
                course_name = next((c.title for c in self.courses if c.code == code), code)
                ltpsc = next((c.ltp for c in self.courses if c.code == code), "")
                faculty = next((c.faculty for c in self.courses if c.code == code), "")
                ws.cell(start_row + i, 4, course_name).border = thin_border
                ws.cell(start_row + i, 5, ltpsc).border = thin_border
                ws.cell(start_row + i, 5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(start_row + i, 6, faculty).border = thin_border
                color_cell = ws.cell(start_row + i, 7, "")
                color_cell.fill = PatternFill(start_color=color_map[code], end_color=color_map[code], fill_type="solid")
                color_cell.border = thin_border
                i += 1

            # tidy borders for timetable region and column widths
            timetable_max_row = len(self.days) + 1
            timetable_max_col = ws.max_column
            for row_cells in ws.iter_rows(min_row=2, max_row=timetable_max_row, min_col=2, max_col=timetable_max_col):
                for cell in row_cells:
                    cell.border = thin_border

            ws.freeze_panes = "B2"
            for col_cells in ws.columns:
                try:
                    column_letter = col_cells[0].column_letter
                except Exception:
                    continue
                max_length = 0
                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Formatted student timetable saved in {filename}")

    
    # --------------------- Full run helper ---------------------
    def run_all_outputs(self, dept_name_prefix="CSE", student_filename=None):
        """
        Generate student timetables (First_Half and Second_Half) and a combined faculty workbook.
        Also writes an unscheduled courses file if any course couldn't be placed.
        """
        if not student_filename:
            student_filename = f"{dept_name_prefix}_timetable.xlsx"

        # reset state for this run
        self.records = []
        self.elective_groups = {}
        self.elective_room_map = {}
        self.unscheduled_list = []

        # write student timetables
        with pd.ExcelWriter(student_filename, engine="openpyxl") as writer:
            # first half (semester_half == "1" or "0")
            self.generate_timetable([c for c in self.courses if c.sem_half in ["1", "0"]], writer, "First_Half")
            # second half (semester_half == "2" or "0")
            self.generate_timetable([c for c in self.courses if c.sem_half in ["2", "0"]], writer, "Second_Half")

        # export unscheduled courses if any
        if self.unscheduled_list:
            unsched_file = f"{dept_name_prefix}_unscheduled_courses.xlsx"
            pd.DataFrame(self.unscheduled_list).to_excel(unsched_file, index=False)
            print(f"Some courses couldn't be scheduled. See '{unsched_file}' for details.")

        # remove default sheets left by pandas writer if any
        wb = load_workbook(student_filename)
        for default in ["Sheet", "Sheet1"]:
            if default in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb[default])
        wb.save(student_filename)

        # compute elective room assignments using final records
        for sheet in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet)

        # format student workbook 
        self.format_student_timetable_with_legend(student_filename)


# --------------------- Script entrypoint ---------------------
if __name__ == "__main__":
    # departments mapping (department_name -> courses csv)
    departments = {
        "CSE-3-A": "data/CSE_3_A_courses.csv",
        "CSE-3-B": "data/CSE_3_B_courses.csv",
        "CSE-1-A": "data/CSE_1_A_courses.csv",
        "CSE-1-B": "data/CSE_1_B_courses.csv",
        "CSE-5-A": "data/CSE_5_A_courses.csv",
        "CSE-5-B": "data/CSE_5_B_courses.csv",
        "7-SEM": "data/DSAI_7_courses.csv",
        "DSAI-3": "data/DSAI_3_courses.csv",
        "ECE-3": "data/ECE_3_courses.csv",
        "DSAI-1": "data/DSAI_1_courses.csv",
        "ECE-1": "data/ECE_1_courses.csv",
        "DSAI-5": "data/DSAI_5_courses.csv",
        "ECE-5": "data/ECE_5_courses.csv",
    }
    rooms_file = "data/rooms.csv"
    slots_file = "data/timeslots.csv"
    global_room_usage = {}

    all_records = []

    # generate per-department timetables
    for dept_name, course_file in departments.items():
        print(f"\nGenerating student timetable for {dept_name}...")
        scheduler = Scheduler(slots_file, course_file, rooms_file, global_room_usage)
        student_file = f"{dept_name}_timetable.xlsx"
        scheduler.run_all_outputs(dept_name_prefix=dept_name, student_filename=student_file)

        # collect scheduled entries and course-room map for a combined faculty book later
        all_records.extend(scheduler.records)
        for k, v in scheduler.course_room_map.items():
            global_room_usage.setdefault("MAPPING", {})[k] = v

    # build combined faculty workbook from all departments
    combined_courses = []
    for dept_name, course_file in departments.items():
        df = pd.read_csv(course_file)
        for _, row in df.iterrows():
            combined_courses.append(Course(row))

    helper = Scheduler(slots_file, departments[list(departments.keys())[0]], rooms_file, global_room_usage)
    helper.courses = combined_courses
    helper.records = all_records
    print("\nAll done. Student timetables and combined faculty timetable generated.")
