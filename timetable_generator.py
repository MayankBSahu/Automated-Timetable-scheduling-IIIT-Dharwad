import pandas as pd
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
FILE_PATH = "CourseCode&Name.csv"
START_DATE = dt.date(2025, 11, 20)
OUTPUT_FILE = "Exam_Timetable.xlsx"

# SHIFT LOGIC (STRATEGY 1)
def get_shift(batch_name: str):
    b = batch_name.upper()
    if "1ST" in b or "3RD" in b:
        return "Morning (10:00 AM – 11:30 AM)"
    else:
        return "Evening (03:00 PM – 04:30 PM)"

# === STEP 1: PARSE THE CUSTOM CSV STRUCTURE ===
courses = {}
current_batch = None
current_courses = []

with open(FILE_PATH, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()

        # Detect batch header
        if line.startswith("BATCH"):
            if current_batch and current_courses:
                courses[current_batch] = current_courses
            current_batch = line.replace("BATCH", "").replace(":", "").strip()
            current_courses = []

        # Detect course line (code, name)
        elif line and "," in line:
            code, name = [x.strip() for x in line.split(",", 1)]
            current_courses.append((code, name))  # store tuple (code, name)

        # Blank line → end of current batch
        elif not line:
            if current_batch and current_courses:
                courses[current_batch] = current_courses
                current_batch, current_courses = None, []

# Save last batch if not yet saved
if current_batch and current_courses:
    courses[current_batch] = current_courses

print(f"✅ Parsed {len(courses)} batches successfully!")

# === STEP 2: GENERATE EXAM DATES (Skipping Weekends) ===
def generate_exam_dates(num_days, start_date):
    dates = []
    d = start_date
    while len(dates) < num_days:
        if d.weekday() < 5:  # Monday–Friday only
            dates.append(d)
        d += dt.timedelta(days=1)
    return dates

max_exams = max(len(c) for c in courses.values())
dates = generate_exam_dates(max_exams + 5, START_DATE)  # extra dates for electives

# === STEP 3: BUILD TIMETABLE DATA ===
records = []

for batch_name, clist in courses.items():

    SHIFT = get_shift(batch_name)  # <-- APPLY 2-SHIFT STRATEGY HERE

    # Separate normal and elective subjects
    normal_courses = [(code, name) for code, name in clist if "ELECTIVE" not in code.upper()]
    elective_courses = [(code, name) for code, name in clist if "ELECTIVE" in code.upper()]

    date_idx = 0
    used_dates = set()

    # Schedule normal courses first
    for code, name in normal_courses:
        while date_idx < len(dates) and dates[date_idx].weekday() >= 5:
            date_idx += 1

        exam_date = dates[date_idx]
        used_dates.add(exam_date)

        records.append({
            "Batch": batch_name,
            "Date": exam_date,
            "Day": exam_date.strftime("%A"),
            "Shift": SHIFT,
            "CourseCode": code,
            "CourseName": name
        })

        date_idx += 1

    # Schedule all electives on one same day
    if elective_courses:
        while date_idx < len(dates) and (dates[date_idx] in used_dates or dates[date_idx].weekday() >= 5):
            date_idx += 1

        elective_date = dates[date_idx]

        for code, name in elective_courses:
            records.append({
                "Batch": batch_name,
                "Date": elective_date,
                "Day": elective_date.strftime("%A"),
                "Shift": SHIFT,
                "CourseCode": code,
                "CourseName": name
            })

        used_dates.add(elective_date)

df = pd.DataFrame(records)

# === STEP 4: SORT AND FORMAT ===
df = df.sort_values(by=["Batch", "Date"]).reset_index(drop=True)
df["Date"] = pd.to_datetime(df["Date"])
df["Date_str"] = df["Date"].dt.strftime("%d-%b-%Y")

# === STEP 5: CREATE EXCEL WORKBOOK ===
wb = Workbook()
ws_master = wb.active
ws_master.title = "Master_Timetable"

# Master sheet header
headers = ["Batch", "Date", "Day", "Shift", "Course Code", "Course Name"]
for col, head in enumerate(headers, start=1):
    ws_master.cell(row=1, column=col, value=head)
    ws_master.cell(row=1, column=col).font = Font(bold=True)
    ws_master.cell(row=1, column=col).alignment = Alignment(horizontal="center")

# Fill master sheet
for i, row in df.iterrows():
    row_data = [row["Batch"], row["Date_str"], row["Day"], row["Shift"], row["CourseCode"], row["CourseName"]]
    for j, val in enumerate(row_data, start=1):
        ws_master.cell(row=i+2, column=j, value=val)
        ws_master.cell(row=i+2, column=j).alignment = Alignment(horizontal="center")

# === STEP 6: INDIVIDUAL SHEETS ===
for batch in df["Batch"].unique():
    sub_df = df[df["Batch"] == batch].sort_values(by="Date")
    ws = wb.create_sheet(title=batch[:30])

    shift_for_sheet = get_shift(batch)  # ensure correct shift title

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Exam Timetable - {batch} ({shift_for_sheet})"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True, size=13)

    # Headers
    for idx, col in enumerate(["Date", "Day", "Course Code", "Course Name"], start=1):
        ws.cell(row=2, column=idx, value=col).font = Font(bold=True)
        ws.cell(row=2, column=idx).alignment = Alignment(horizontal="center")

    # Data rows
    for i, r in enumerate(sub_df.itertuples(), start=3):
        ws.cell(row=i, column=1, value=r.Date_str)
        ws.cell(row=i, column=2, value=r.Day)
        ws.cell(row=i, column=3, value=r.CourseCode)
        ws.cell(row=i, column=4, value=r.CourseName)
        for c in range(1, 5):
            ws.cell(row=i, column=c).alignment = Alignment(horizontal="center")

    # Borders + Auto width
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=2, max_row=i, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

# Save Excel
wb.save(OUTPUT_FILE)
print(f"🎯 Timetable created successfully → {OUTPUT_FILE}")
