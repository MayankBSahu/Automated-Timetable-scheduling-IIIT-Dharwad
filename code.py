

import pandas as pd
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import sys
import math
from pathlib import Path
from collections import OrderedDict

def get_user_date(prompt):
    while True:
        s = input(prompt + " (DD-MM-YYYY): ").strip()
        try:
            return dt.datetime.strptime(s, "%d-%m-%Y").date()
        except Exception:
            print("Invalid format. Use DD-MM-YYYY.")

print("=== Exam Timetable Generator (final) ===")
START_DATE = get_user_date("Enter exam START date")
END_DATE   = get_user_date("Enter exam END date")
if END_DATE < START_DATE:
    print("End date is before start date. Exiting.")
    sys.exit(1)


SLOTS = ["Morning (10:00 AM – 11:30 AM)", "Afternoon (02:00 PM – 03:30 PM)"]

COURSE_FILE = "FINAL_EXCEL.csv"
ROOM_FILE = "rooms.csv"
OUTPUT_FILE = "Exam_Timetable_Final.xlsx"


def safe_int(x):
    try:
        if pd.isna(x):
            return None
        s = str(x).replace(",", "").strip()
        if s in ("", "-", "NA", "N/A", "nan"):
            return None
        return int(float(s))
    except:
        return None

def generate_weekdays(start, end):
    d = start
    res = []
    while d <= end:
        if d.weekday() < 5:
            res.append(d)
        d += dt.timedelta(days=1)
    return res

if not Path(COURSE_FILE).exists():
    print(f"Missing file: {COURSE_FILE}")
    sys.exit(1)
if not Path(ROOM_FILE).exists():
    print(f"Missing file: {ROOM_FILE}")
    sys.exit(1)

df = pd.read_csv(COURSE_FILE)
rooms_df = pd.read_csv(ROOM_FILE)

df.columns = [c.strip().lower() for c in df.columns]
rooms_df.columns = [c.strip().lower() for c in rooms_df.columns]

req_cols = ["course code", "batch_real", "no. of students"]
missing = [c for c in req_cols if c not in df.columns]
if missing:
    print("Warning: FINAL_EXCEL.csv missing expected columns:", missing)

SPLIT_MAP = {
    "all 1st-year": ["1csea", "1cseb", "1ece", "1dsai"],
    "all 1st year": ["1csea", "1cseb", "1ece", "1dsai"],
    "all-1st-year": ["1csea", "1cseb", "1ece", "1dsai"],
    "all-2nd year": ["2cse-a", "2cse-b", "2ece", "2dsai"],
    "all-2nd-year": ["2cse-a", "2cse-b", "2ece", "2dsai"],
    "all-2nd year": ["2cse-a", "2cse-b", "2ece", "2dsai"],
    "all-3rd year": ["3csea", "3cseb", "3ece", "3dsai"],
    "all-3rd-year": ["3csea", "3cseb", "3ece", "3dsai"],
    "all 3rd year": ["3csea", "3cseb", "3ece", "3dsai"],
    "all-4th year": ["4cse", "4ece", "4dsai"],
    "all-4th-year": ["4cse", "4ece", "4dsai"],
    "all 4th year": ["4cse", "4ece", "4dsai"],
}

expanded_rows = []  
for _, row in df.iterrows():
    course = str(row.get("course code", "")).strip()
    batch_raw = str(row.get("batch_real", "")).strip()
    batch_key = batch_raw.lower()
    students = safe_int(row.get("no. of students")) or 0

    if batch_key in SPLIT_MAP:
        targets = SPLIT_MAP[batch_key]
        n = len(targets)
        if n == 0:
            continue
        base, rem = divmod(students, n)
       
        for i, t in enumerate(targets):
            s = base + (1 if i < rem else 0)
            expanded_rows.append({"batch": t.upper(), "course": course, "students": int(s)})
    else:
      
        expanded_rows.append({"batch": batch_raw, "course": course, "students": int(students)})

if not expanded_rows:
    print("No courses found to schedule after expansion. Exiting.")
    sys.exit(1)

courses_by_batch = OrderedDict()
for r in expanded_rows:
    b = r["batch"]
    courses_by_batch.setdefault(b, []).append({"course": r["course"], "students": r["students"]})

room_id_col = None
cap_col = None
for c in rooms_df.columns:
    if c in ("room", "room_id", "roomid", "room id"):
        room_id_col = c
    if c in ("capacity", "cap", "seats", "seat"):
        cap_col = c

if room_id_col is None:
    room_id_col = rooms_df.columns[0]
if cap_col is None:
   
    for c in rooms_df.columns:
        if rooms_df[c].apply(lambda x: safe_int(x) is not None).sum() > 0:
            cap_col = c
            break

if cap_col is None:
    print("Error: could not find capacity column in rooms.csv")
    sys.exit(1)

rooms_list = []
for _, r in rooms_df.iterrows():
    room_name = str(r[room_id_col]).strip()
    cap_raw = safe_int(r[cap_col])
    if cap_raw is None or cap_raw <= 0:
        continue
    usable = cap_raw // 2  
    per_course_quota = usable // 2  
    if per_course_quota <= 0:
        continue
    rooms_list.append({"room": room_name, "per_course_quota": int(per_course_quota)})

if not rooms_list:
    print("No usable rooms after applying half-capacity + per-course split. Exiting.")
    sys.exit(1)

rooms_list.sort(key=lambda x: x["per_course_quota"], reverse=True)


dates = generate_weekdays(START_DATE, END_DATE)
if not dates:
    print("No weekdays in given date range. Exiting.")
    sys.exit(1)

date_slots = [(d, slot) for d in dates for slot in SLOTS]

room_availability = {}
for d, slot in date_slots:
    key = (d.strftime("%Y-%m-%d"), slot)
    room_availability[key] = {
        r["room"]: {"assigned_courses": 0, "remaining_quota": r["per_course_quota"]}
        for r in rooms_list
    }

def allocate_rooms_for_course(students_needed, date_str, slot):
    
    if students_needed <= 0:
        return []

    state = room_availability.get((date_str, slot))
    if state is None:
        return []

    empty_rooms = [(room, info["remaining_quota"]) for room, info in state.items() if info["assigned_courses"] == 0 and info["remaining_quota"] > 0]
    one_course_rooms = [(room, info["remaining_quota"]) for room, info in state.items() if info["assigned_courses"] == 1 and info["remaining_quota"] > 0]

   
    candidates = empty_rooms + one_course_rooms
    candidates.sort(key=lambda x: x[1], reverse=True)  

    assigned = []
    remaining = students_needed

    for room, cap in candidates:
        if remaining <= 0:
            break
        info = state[room]
        if info["assigned_courses"] >= 2 or info["remaining_quota"] <= 0:
            continue
        take = min(info["remaining_quota"], remaining)
        assigned.append((room, int(take)))
        info["remaining_quota"] -= take
        info["assigned_courses"] += 1
        remaining -= take

    return assigned

records = []
warning_short_capacity = False

for batch, course_list in courses_by_batch.items():
    used_calendar_dates = set()  

    for entry in course_list:
        course_code = entry["course"]
        students = int(entry["students"])

        scheduled = False

        for (d, slot) in date_slots:
            date_str = d.strftime("%Y-%m-%d")
            if date_str in used_calendar_dates:
                continue

            total_avail = sum(info["remaining_quota"] for info in room_availability[(date_str, slot)].values() if info["assigned_courses"] < 2)
            if total_avail <= 0:
                continue

            assigned = allocate_rooms_for_course(students, date_str, slot)
            if sum(a for _, a in assigned) >= students:
            
                used_calendar_dates.add(date_str)
                records.append({
                    "Batch": batch,
                    "Date": d,
                    "Date_str": d.strftime("%d-%b-%Y"),
                    "Day": d.strftime("%A"),
                    "Slot": slot,
                    "Course": course_code,
                    "Students": students,
                    "Rooms": "; ".join(f"{r} ({c})" for r, c in assigned)
                })
                scheduled = True
                break
            elif assigned:
                
                warning_short_capacity = True
                used_calendar_dates.add(date_str)
                records.append({
                    "Batch": batch,
                    "Date": d,
                    "Date_str": d.strftime("%d-%b-%Y"),
                    "Day": d.strftime("%A"),
                    "Slot": slot,
                    "Course": course_code,
                    "Students": students,
                    "Rooms": "; ".join(f"{r} ({c})" for r, c in assigned) + " (PARTIAL)"
                })
                scheduled = True
                break
           

        if not scheduled:
            
            best_key = None
            best_total = 0
            for (d, slot) in date_slots:
                date_str = d.strftime("%Y-%m-%d")
                total_avail = sum(info["remaining_quota"] for info in room_availability[(date_str, slot)].values() if info["assigned_courses"] < 2)
                if total_avail > best_total:
                    best_total = total_avail
                    best_key = (d, slot)
            if best_key is None or best_total == 0:
                
                records.append({
                    "Batch": batch,
                    "Date": dates[0],
                    "Date_str": dates[0].strftime("%d-%b-%Y"),
                    "Day": dates[0].strftime("%A"),
                    "Slot": SLOTS[0],
                    "Course": course_code,
                    "Students": students,
                    "Rooms": ""
                })
                warning_short_capacity = True
            else:
                d_best, slot_best = best_key
                date_str = d_best.strftime("%Y-%m-%d")
                assigned = allocate_rooms_for_course(students, date_str, slot_best)
                if sum(a for _, a in assigned) < students:
                    warning_short_capacity = True
                records.append({
                    "Batch": batch,
                    "Date": d_best,
                    "Date_str": d_best.strftime("%d-%b-%Y"),
                    "Day": d_best.strftime("%A"),
                    "Slot": slot_best,
                    "Course": course_code,
                    "Students": students,
                    "Rooms": "; ".join(f"{r} ({c})" for r, c in assigned) + (" (PARTIAL)" if sum(a for _,a in assigned) < students else "")
                })


out_df = pd.DataFrame(records)
if out_df.empty:
    print("No records scheduled. Exiting.")
    sys.exit(1)

out_df = out_df.sort_values(by=["Batch", "Date", "Slot", "Course"]).reset_index(drop=True)

wb = Workbook()
ws = wb.active
ws.title = "Master_Timetable"

headers = ["Batch", "Date", "Day", "Slot", "Course", "Students", "Rooms"]


for i, h in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    ws.cell(row=1, column=i).alignment = Alignment(horizontal="center")


for r_idx, row in out_df.iterrows():
    vals = [row["Batch"], row["Date_str"], row["Day"], row["Slot"], row["Course"], row["Students"], row["Rooms"]]
    for c_idx, v in enumerate(vals, start=1):
        ws.cell(row=r_idx + 2, column=c_idx, value=v)
        ws.cell(row=r_idx + 2, column=c_idx).alignment = Alignment(horizontal="center")


for batch in out_df["Batch"].unique():
    sub = out_df[out_df["Batch"] == batch].reset_index(drop=True)
    sheet_title = str(batch)[:31]
    ws2 = wb.create_sheet(title=sheet_title)
    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"Exam Timetable - {batch}"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A1"].alignment = Alignment(horizontal="center")

    for i, h in enumerate(headers, start=1):
        ws2.cell(row=2, column=i, value=h).font = Font(bold=True)
        ws2.cell(row=2, column=i).alignment = Alignment(horizontal="center")

    for i, r in sub.iterrows():
        ws2.cell(row=i + 3, column=1, value=r["Batch"])
        ws2.cell(row=i + 3, column=2, value=r["Date_str"])
        ws2.cell(row=i + 3, column=3, value=r["Day"])
        ws2.cell(row=i + 3, column=4, value=r["Slot"])
        ws2.cell(row=i + 3, column=5, value=r["Course"])
        ws2.cell(row=i + 3, column=6, value=r["Students"])
        ws2.cell(row=i + 3, column=7, value=r["Rooms"])
        for c in range(1, 8):
            ws2.cell(row=i + 3, column=c).alignment = Alignment(horizontal="center")


for col_cells in ws.columns:
    max_len = 0
    col = col_cells[0].column_letter
    for cell in col_cells:
        try:
            v = str(cell.value or "")
            if len(v) > max_len:
                max_len = len(v)
        except:
            pass
    ws.column_dimensions[col].width = min(max(10, max_len + 2), 60)


out_path = Path(OUTPUT_FILE)
i = 1
base = out_path.stem
ext = out_path.suffix or ".xlsx"
while True:
    try:
        wb.save(out_path)
        break
    except PermissionError:
        out_path = Path(f"{base}_{i}{ext}")
        i += 1
print(f"\n🎯 Timetable generated → {out_path}")

if warning_short_capacity:
    print("⚠ Warning: some courses were only partially accommodated due to limited room capacity. Check '(PARTIAL)' tags in the Rooms column.")

print("Done — let me know if you want per-room seating lists, printed timetables, or further rules.")
